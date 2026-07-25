import csv
import sys
import os
import subprocess

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def build_test_cases():
    columns = [
        "Test Case ID", "Module", "Sub Module", "Scenario ID", "Scenario description",
        "Tester Name", "Date", "Test case description", "Test Steps", "Test Data",
        "Type of Test cases", "Expected results", "Date of execution", "Executed By",
        "Test Status", "Actual Result", "Defect ID"
    ]

    raw_cases = [
        # 1. CRM Dashboard
        ("TC_CRM_001", "CRM", "Dashboard", "SC_CRM_01", "Dashboard Metrics Overview",
         "Verify that all KPI summary cards (Total Leads, New Leads Today, Hot Leads, Conversion Rate, Pending Follow-ups) display accurate counts",
         "1. Log in as Admin/Counselor\n2. Navigate to CRM > Dashboard\n3. Check the numerical values on KPI cards against actual database/list records",
         "None", "Positive", "All KPI cards should display exact and updated counts corresponding to real-time lead data."),
        
        ("TC_CRM_002", "CRM", "Dashboard", "SC_CRM_01", "Dashboard Metrics Overview",
         "Verify Lead Trend line chart (Last 7 Days) renders correctly and reflects accurate daily lead counts",
         "1. Navigate to CRM > Dashboard\n2. Observe the 'Lead Trend' line chart\n3. Hover over data points for each day",
         "None", "Positive", "Chart should render without visual overlapping and tooltip should display correct daily lead count."),
        
        ("TC_CRM_003", "CRM", "Dashboard", "SC_CRM_01", "Dashboard Metrics Overview",
         "Verify Lead Pipeline Distribution pie chart and Lead Sources bar chart loading and interactivity",
         "1. Navigate to CRM Dashboard\n2. Check Pie Chart for Stage distribution\n3. Check Bar Chart for Lead Sources\n4. Click/Hover on chart segments",
         "None", "Positive", "Charts should accurately categorize leads by Stage and Source with interactive tooltips working smoothly."),
        
        ("TC_CRM_004", "CRM", "Dashboard", "SC_CRM_01", "Dashboard Metrics Overview",
         "Verify 'Recent Leads (Last 5)' list displays the 5 most recently created/updated leads",
         "1. Navigate to CRM Dashboard\n2. Check the 'Recent Leads' section\n3. Verify order and lead details (Name, Stage, Source)",
         "None", "Positive", "The list should display exactly 5 most recent leads sorted in descending order of creation/update timestamp."),
        
        ("TC_CRM_005", "CRM", "Dashboard", "SC_CRM_01", "Dashboard Metrics Overview",
         "Verify Quick Action buttons on Dashboard (Add New Lead, View Kanban, Schedule Follow-up, View Analytics) redirect to correct modules",
         "1. On CRM Dashboard click 'Add New Lead'\n2. Verify modal/page opens\n3. Repeat for 'View Kanban' and 'Schedule Follow-up'",
         "None", "Positive", "Each quick action button should instantly open the corresponding modal or redirect to the correct target URL without errors."),

        # 2. Lead Management
        ("TC_CRM_006", "CRM", "Lead Management", "SC_CRM_02", "Add New Lead",
         "Verify adding a new lead with all mandatory fields filled validly",
         "1. Navigate to CRM > Leads\n2. Click 'Add New Lead' button\n3. Enter Full Name, Mobile Number, Email, Course Interest, and Source\n4. Click 'Save Lead'",
         "Name: Rahul Sharma\nMobile: 9876543210\nEmail: rahul@example.com\nCourse: Full Stack Dev\nSource: Website", "Positive", "New lead should be created successfully with default stage as 'New' and a success toast notification should appear."),
        
        ("TC_CRM_007", "CRM", "Lead Management", "SC_CRM_02", "Add New Lead",
         "Verify validation errors when mandatory fields (Name, Mobile, Source) are left blank",
         "1. Click 'Add New Lead'\n2. Leave Name and Mobile fields empty\n3. Click 'Save Lead'",
         "Name: (blank)\nMobile: (blank)", "Negative", "System should prevent form submission and display clear red validation messages below empty mandatory fields."),
        
        ("TC_CRM_008", "CRM", "Lead Management", "SC_CRM_02", "Add New Lead",
         "Verify Mobile Number validation for invalid format or incorrect digit count",
         "1. Click 'Add New Lead'\n2. Enter 5-digit number or alphabetic string in Mobile field\n3. Click 'Save Lead'",
         "Mobile: 12345 or 98765abcd", "Negative", "System should reject invalid phone numbers and show error 'Please enter a valid 10-digit mobile number'."),
        
        ("TC_CRM_009", "CRM", "Lead Management", "SC_CRM_02", "Add New Lead",
         "Verify Email validation for improper format during lead creation",
         "1. Click 'Add New Lead'\n2. Enter invalid email string\n3. Click 'Save Lead'",
         "Email: rahul.sharma@com or rahul@", "Negative", "System should display validation error 'Invalid email address format' and prevent saving."),
        
        ("TC_CRM_010", "CRM", "Lead Management", "SC_CRM_02", "Add New Lead",
         "Verify duplicate lead prevention when entering an existing Mobile Number or Email",
         "1. Click 'Add New Lead'\n2. Enter Mobile or Email already assigned to an existing lead\n3. Click 'Save Lead'",
         "Mobile: 9876543210 (existing)\nEmail: rahul@example.com (existing)", "Negative", "System should warn 'Lead with this Mobile/Email already exists' and offer an option to view existing lead."),
        
        ("TC_CRM_011", "CRM", "Lead Management", "SC_CRM_02", "Add New Lead",
         "Verify form reset/cancel button clears all entered data in Lead Creation modal",
         "1. Click 'Add New Lead'\n2. Fill all fields with sample data\n3. Click 'Cancel' or 'Reset'\n4. Reopen modal",
         "Sample test data", "Positive", "Form fields should be completely reset to default blank state upon reopening."),
        
        ("TC_CRM_012", "CRM", "Lead Management", "SC_CRM_03", "Lead List View & Filters",
         "Verify Lead List view pagination controls (Next, Previous, Page Numbers)",
         "1. Navigate to CRM > Leads list\n2. Change page size (e.g., 10, 25, 50)\n3. Click 'Next' and 'Previous' buttons",
         "None", "Positive", "Table should display correct number of rows per page and navigate smoothly without duplicate or missing records."),
        
        ("TC_CRM_013", "CRM", "Lead Management", "SC_CRM_03", "Lead List View & Filters",
         "Verify global search bar functionality in Lead List (Search by Name, Email, or Phone)",
         "1. In Leads list search bar enter a lead's Name, partial Phone, or Email\n2. Press Enter or wait for debounce",
         "Search Query: 'Rahul' or '98765'", "Positive", "List should dynamically filter to show only leads matching the search term across Name, Email, and Phone columns."),
        
        ("TC_CRM_014", "CRM", "Lead Management", "SC_CRM_03", "Lead List View & Filters",
         "Verify multi-parameter filtering by Lead Stage, Course Interest, and Source",
         "1. Open Filter dropdown in Leads list\n2. Select Stage = 'Qualified' and Source = 'Instagram'\n3. Apply filter",
         "Stage: Qualified\nSource: Instagram", "Positive", "Table should display only those leads that satisfy all selected filter criteria simultaneously."),
        
        ("TC_CRM_015", "CRM", "Lead Management", "SC_CRM_03", "Lead List View & Filters",
         "Verify bulk selection and export functionality from Lead List view",
         "1. Check select-all checkbox or select 3 individual leads\n2. Click 'Export Selected' or 'Bulk Action' dropdown",
         "3 Selected Leads", "Positive", "System should generate and download an Excel/CSV file containing exact data of only the selected leads."),

        # 3. Kanban Board
        ("TC_CRM_016", "CRM", "Kanban Board", "SC_CRM_04", "Kanban Stage Pipeline",
         "Verify Kanban Board renders all 6 standard stages (New, Contacted, Qualified, Negotiation, Converted, Lost)",
         "1. Navigate to CRM > Kanban\n2. Inspect column headers and structure",
         "None", "Positive", "All 6 columns should be displayed in correct sequence with distinct color coding and accurate lead count badges."),
        
        ("TC_CRM_017", "CRM", "Kanban Board", "SC_CRM_04", "Kanban Stage Pipeline",
         "Verify drag-and-drop functionality to move a lead from 'New' to 'Contacted' stage",
         "1. On Kanban board click and hold a lead card in 'New' column\n2. Drag and drop it into 'Contacted' column\n3. Release mouse",
         "Lead Card in 'New' stage", "Positive", "Card should smoothly snap into 'Contacted' column; column counts should update immediately (+1 in Contacted, -1 in New)."),
        
        ("TC_CRM_018", "CRM", "Kanban Board", "SC_CRM_04", "Kanban Stage Pipeline",
         "Verify persistence of stage change after drag-and-drop (Auto-save verification)",
         "1. Drag lead from 'Contacted' to 'Qualified' on Kanban board\n2. Refresh browser page (F5)\n3. Verify lead position",
         "Lead Card moved to 'Qualified'", "Positive", "The lead must remain in the new 'Qualified' column after page refresh without requiring manual save button click."),
        
        ("TC_CRM_019", "CRM", "Kanban Board", "SC_CRM_04", "Kanban Stage Pipeline",
         "Verify Quick Action buttons directly on Kanban Lead Cards (Call, WhatsApp, Email, View)",
         "1. Locate a lead card on Kanban board\n2. Click on 'WhatsApp' icon\n3. Click on 'Call' (tel:) icon\n4. Click 'View' icon",
         "Lead card with valid phone/email", "Positive", "WhatsApp link should open web.whatsapp.com with pre-filled number; Call should trigger phone handler; View should open Lead Detail page."),
        
        ("TC_CRM_020", "CRM", "Kanban Board", "SC_CRM_04", "Kanban Stage Pipeline",
         "Verify Kanban column scrolling when a stage contains a large number of leads (20+ leads)",
         "1. Navigate to Kanban board\n2. Locate a column with 20+ leads\n3. Scroll vertically within that specific column",
         "Column with 20+ leads", "Positive", "Only the specific column should scroll vertically while keeping the column header and overall board layout fixed."),
        
        ("TC_CRM_021", "CRM", "Kanban Board", "SC_CRM_04", "Kanban Stage Pipeline",
         "Verify filter application on Kanban Board view (Filter by Counselor or Source)",
         "1. On Kanban page select Filter -> Source: 'Google Ads'\n2. Observe cards across all columns",
         "Filter: Source = Google Ads", "Positive", "Kanban columns should dynamically hide non-matching cards and recalculate stage badges for visible cards only."),
        
        ("TC_CRM_022", "CRM", "Kanban Board", "SC_CRM_04", "Kanban Stage Pipeline",
         "Verify error handling if network fails during drag-and-drop stage transition",
         "1. Simulate network offline / disconnect\n2. Drag lead from 'New' to 'Lost' on Kanban",
         "Offline state simulation", "Negative", "System should show error toast 'Failed to update lead stage' and revert the lead card back to its original column."),

        # 4. Lead Details
        ("TC_CRM_023", "CRM", "Lead Details", "SC_CRM_05", "Lead 360 Profile & Timeline",
         "Verify Lead Detail Page loads complete profile information and avatar correctly",
         "1. Click on any Lead Name from List or Kanban\n2. Inspect Profile header, Contact info, Course interest, and Source badges",
         "Valid Lead ID", "Positive", "All lead attributes should match database records exactly; avatar should render initials or uploaded photo cleanly."),
        
        ("TC_CRM_024", "CRM", "Lead Details", "SC_CRM_05", "Lead 360 Profile & Timeline",
         "Verify inline editing of lead contact details (Mobile, Email, Address) on Detail Page",
         "1. On Lead Detail page click 'Edit' icon next to Email/Phone\n2. Modify the value\n3. Click Save/Checkmark icon",
         "Updated Email: rahul.new@example.com", "Positive", "Field should update in-place without full page reload and display a success notification toast."),
        
        ("TC_CRM_025", "CRM", "Lead Details", "SC_CRM_05", "Lead 360 Profile & Timeline",
         "Verify Activities Timeline tab displays chronological audit log of all lead interactions",
         "1. On Lead Detail page switch to 'Activities Timeline' tab\n2. Review log entries for creation, stage changes, and follow-ups",
         "Lead with past interactions", "Positive", "Timeline should display all historical events in reverse chronological order with exact timestamps and user names."),
        
        ("TC_CRM_026", "CRM", "Lead Details", "SC_CRM_05", "Lead 360 Profile & Timeline",
         "Verify adding a new internal Note in the 'Notes' tab",
         "1. Go to 'Notes' tab on Lead Detail page\n2. Enter text in 'Add a note...' textarea\n3. Click 'Post Note' button",
         "Note Text: 'Student requested fee discount details'", "Positive", "Note should be saved instantly and appear at the top of the notes list with timestamp and author name."),
        
        ("TC_CRM_027", "CRM", "Lead Details", "SC_CRM_05", "Lead 360 Profile & Timeline",
         "Verify validation when trying to post an empty or whitespace-only Note",
         "1. Go to 'Notes' tab\n2. Leave textarea empty or type only spaces\n3. Try clicking 'Post Note'",
         "Note Text: \"   \" (spaces)", "Negative", "The 'Post Note' button should be disabled or show validation warning 'Note content cannot be empty'."),
        
        ("TC_CRM_028", "CRM", "Lead Details", "SC_CRM_05", "Lead 360 Profile & Timeline",
         "Verify 'Follow-ups History' tab lists all past and upcoming scheduled follow-ups",
         "1. Go to 'Follow-ups History' tab on Lead Detail page\n2. Check list of scheduled, completed, and overdue follow-ups",
         "Lead with scheduled follow-ups", "Positive", "All follow-up records associated with this lead should display with date, time, remarks, and current status badge."),
        
        ("TC_CRM_029", "CRM", "Lead Details", "SC_CRM_05", "Lead 360 Profile & Timeline",
         "Verify Quick Actions sidebar buttons on Detail Page (Call, WhatsApp, Email, Schedule Follow-up)",
         "1. On Lead Detail page locate right sidebar\n2. Test each quick action button individually",
         "None", "Positive", "Each sidebar button should initiate the respective communication channel or open the schedule modal seamlessly."),

        # 5. Follow-ups Management
        ("TC_CRM_030", "CRM", "Follow-ups", "SC_CRM_06", "Schedule & Manage Follow-ups",
         "Verify scheduling a new follow-up with valid future Date, Time, and Remarks",
         "1. Click 'Schedule Follow-up' from Dashboard or Lead Detail\n2. Select Date (tomorrow), Time (11:00 AM), Type (Call), Remarks\n3. Click Save",
         "Date: Tomorrow\nTime: 11:00 AM\nType: Phone Call\nRemarks: Discuss batch timing", "Positive", "Follow-up should be scheduled successfully and appear in 'Today/Upcoming Follow-ups' list and lead's timeline."),
        
        ("TC_CRM_031", "CRM", "Follow-ups", "SC_CRM_06", "Schedule & Manage Follow-ups",
         "Verify validation error when attempting to schedule a follow-up in the past",
         "1. Open 'Schedule Follow-up' modal\n2. Select a Date and Time from yesterday\n3. Click Save",
         "Date: Yesterday's date\nTime: 10:00 AM", "Negative", "System should reject past date/time and display error 'Follow-up date and time must be in the future'."),
        
        ("TC_CRM_032", "CRM", "Follow-ups", "SC_CRM_06", "Schedule & Manage Follow-ups",
         "Verify updating follow-up status from 'Pending' to 'Completed' with completion notes",
         "1. Navigate to Dashboard or Lead Follow-ups tab\n2. Click 'Mark Complete' on a pending follow-up\n3. Enter completion remarks\n4. Save",
         "Status: Completed\nRemarks: Spoke to student, interested in weekend batch", "Positive", "Follow-up status should change to 'Completed' (green badge) and action should be logged in Activities Timeline."),
        
        ("TC_CRM_033", "CRM", "Follow-ups", "SC_CRM_06", "Schedule & Manage Follow-ups",
         "Verify rescheduling an existing pending follow-up to a new date and time",
         "1. Locate a pending follow-up\n2. Click 'Reschedule' icon\n3. Pick a new future date/time\n4. Enter reschedule reason\n5. Click Update",
         "New Date: +3 days\nReason: Student busy today", "Positive", "Follow-up date/time should update successfully and status should remain 'Pending' with new schedule displayed."),
        
        ("TC_CRM_034", "CRM", "Follow-ups", "SC_CRM_06", "Schedule & Manage Follow-ups",
         "Verify Overdue Follow-ups highlighting and badge indicators",
         "1. Check Dashboard or Follow-ups list for follow-ups where scheduled date/time has passed without completion\n2. Observe visual styling",
         "Overdue follow-up records", "Positive", "Overdue follow-ups should be prominently highlighted in red/orange with an 'Overdue' warning badge and alert count."),
        
        ("TC_CRM_035", "CRM", "Follow-ups", "SC_CRM_06", "Schedule & Manage Follow-ups",
         "Verify canceling/deleting a scheduled follow-up",
         "1. Locate a scheduled follow-up\n2. Click 'Cancel/Delete' icon\n3. Confirm deletion in prompt",
         "Confirmation dialog", "Positive", "Follow-up should be removed from active lists and marked as cancelled in the lead's historical audit log."),

        # 6. Lead Conversion & Closure
        ("TC_CRM_036", "CRM", "Lead Conversion", "SC_CRM_07", "Convert & Close Leads",
         "Verify converting a Qualified Lead into a Student/Admission successfully",
         "1. Open a Qualified lead on Detail page\n2. Click 'Convert to Student' button\n3. Verify pre-filled data on Admission form\n4. Confirm conversion",
         "Lead in 'Qualified' stage", "Positive", "Lead status should change to 'Converted', a new Student profile should be created with pre-filled contact data, and link established."),
        
        ("TC_CRM_037", "CRM", "Lead Conversion", "SC_CRM_07", "Convert & Close Leads",
         "Verify that converting a lead locks or disables further stage modifications on Kanban",
         "1. Convert a lead to Student ('Converted' stage)\n2. Navigate to Kanban board\n3. Attempt to drag the converted lead card to 'New' or 'Contacted'",
         "Converted Lead Card", "Negative", "Converted leads should be locked from backward stage movement; dragging should be disabled or show warning toast."),
        
        ("TC_CRM_038", "CRM", "Lead Conversion", "SC_CRM_07", "Convert & Close Leads",
         "Verify marking a Lead as 'Lost' with mandatory Loss Reason selection",
         "1. On Lead Detail or Kanban click 'Mark as Lost'\n2. Try submitting without selecting reason\n3. Select Reason (e.g., 'High Fees', 'Joined Competitor')\n4. Save",
         "Loss Reason: 'Joined Competitor'\nRemarks: Opted for online course", "Positive", "System should require selecting a Loss Reason before marking as Lost; lead stage should update to 'Lost' with reason recorded."),
        
        ("TC_CRM_039", "CRM", "Lead Conversion", "SC_CRM_07", "Convert & Close Leads",
         "Verify reopening a 'Lost' lead back to 'Contacted' or 'Negotiation' stage",
         "1. Open a lead in 'Lost' stage\n2. Click 'Reopen Lead' or change stage dropdown to 'Contacted'\n3. Enter reopening reason\n4. Save",
         "Reopen Reason: Student reached out again for new batch", "Positive", "Lead should be restored to active pipeline ('Contacted') and removed from Lost analytics buckets."),
        
        ("TC_CRM_040", "CRM", "Lead Conversion", "SC_CRM_07", "Convert & Close Leads",
         "Verify conversion analytics and win/loss ratios update immediately on CRM Dashboard",
         "1. Convert 1 lead and mark 1 lead as Lost\n2. Go to CRM > Dashboard\n3. Check Conversion Rate % and Stage distribution charts",
         "1 Converted Lead, 1 Lost Lead", "Positive", "Dashboard KPI metrics and conversion percentage should immediately recalculate to reflect the newly closed deals."),

        # 7. Security & Export
        ("TC_CRM_041", "CRM", "Security & Export", "SC_CRM_08", "Permissions & Data Management",
         "Verify Role-Based Access Control (RBAC) - Counselor visibility restriction",
         "1. Log in as a restricted Counselor user\n2. Navigate to CRM > Leads\n3. Check if user can only see assigned leads vs all company leads",
         "Counselor Role Login", "Positive", "Counselor should only view and manage leads assigned to them (if restricted mode is enabled in CRM settings)."),
        
        ("TC_CRM_042", "CRM", "Security & Export", "SC_CRM_08", "Permissions & Data Management",
         "Verify Counselor cannot delete leads if Delete permission is disabled for their role",
         "1. Log in as Counselor\n2. Open Lead Detail page or List view\n3. Look for 'Delete Lead' option or try calling delete API",
         "Counselor Role Login", "Negative", "The 'Delete' button should be hidden/disabled for Counselor role, and direct API delete request should return 403 Forbidden."),
        
        ("TC_CRM_043", "CRM", "Security & Export", "SC_CRM_08", "Permissions & Data Management",
         "Verify exporting full CRM leads database to Excel/CSV format",
         "1. Log in as Admin\n2. Navigate to CRM > Leads\n3. Click 'Export All Leads' button\n4. Open downloaded spreadsheet",
         "Admin Role Login", "Positive", "System should export a valid .csv/.xlsx file containing all lead records with accurate headers and formatted dates."),
        
        ("TC_CRM_044", "CRM", "Security & Export", "SC_CRM_08", "Permissions & Data Management",
         "Verify importing leads via CSV upload with field mapping",
         "1. Navigate to CRM > Leads > Import\n2. Upload valid CSV with 10 test leads\n3. Map columns (Name, Phone, Email)\n4. Execute Import",
         "Sample CSV with 10 rows", "Positive", "All 10 leads should be imported into 'New' stage without data truncation; summary report should show 10 successful imports."),
        
        ("TC_CRM_045", "CRM", "Security & Export", "SC_CRM_08", "Permissions & Data Management",
         "Verify import error reporting when uploaded CSV contains invalid phone numbers or duplicate emails",
         "1. Upload CSV where row 3 has invalid phone and row 5 has duplicate email\n2. Execute Import\n3. Review Import Summary log",
         "CSV with intentional errors in rows 3 and 5", "Negative", "System should import valid rows while skipping invalid rows and generating an error log detailing exact row numbers and failure reasons."),

        # 8. Document Upload & Management
        ("TC_CRM_046", "CRM", "Document Upload", "SC_CRM_09", "Document Upload & Management",
         "Verify uploading a valid document (PDF/JPG/PNG under 5MB) on Lead Profile",
         "1. Open Lead Detail page\n2. Locate Documents section / Upload Document button\n3. Select a valid 2MB PDF file (e.g. ID proof)\n4. Click Upload",
         "File: Student_ID_Proof.pdf (2MB)", "Positive", "Document should upload successfully and appear in the documents list with filename, upload date, and file size."),
        
        ("TC_CRM_047", "CRM", "Document Upload", "SC_CRM_09", "Document Upload & Management",
         "Verify file size and format validation during document upload",
         "1. On Lead Detail page try uploading an unsupported file (.exe, .bat) or a file >10MB\n2. Observe validation response",
         "File: setup.exe or large_video.mp4 (15MB)", "Negative", "System should reject the upload with an error message 'Unsupported file format' or 'File size exceeds 5MB limit'."),
        
        ("TC_CRM_048", "CRM", "Document Upload", "SC_CRM_09", "Document Upload & Management",
         "Verify downloading and previewing uploaded documents",
         "1. Locate an uploaded document on Lead Detail page\n2. Click 'Download' or 'Preview' icon",
         "Uploaded document record", "Positive", "Document should open cleanly in a preview modal or trigger a browser download without file corruption."),

        # 9. Dedicated CRM Analytics
        ("TC_CRM_049", "CRM", "Analytics Page", "SC_CRM_10", "Dedicated Analytics (/crm/analytics)",
         "Verify Lead Stage Funnel chart on Analytics page accurately represents pipeline drop-off",
         "1. Navigate to CRM > Analytics (/crm/analytics)\n2. Inspect Lead Stage Funnel chart\n3. Compare stage counts against database",
         "Real-time pipeline data", "Positive", "Funnel chart should display accurate step-by-step conversion counts and drop-off percentages from New to Converted."),
        
        ("TC_CRM_050", "CRM", "Analytics Page", "SC_CRM_10", "Dedicated Analytics (/crm/analytics)",
         "Verify Counsellor Performance table calculations (Assigned Leads, Contacted %, Conversion Rate)",
         "1. On CRM Analytics page scroll to 'Counsellor Performance' table\n2. Verify metrics for each individual counselor",
         "Multi-counselor assignment data", "Positive", "Table should display exact assigned counts, percentage contacted within SLA, and conversion ratios per counselor."),
        
        ("TC_CRM_051", "CRM", "Analytics Page", "SC_CRM_10", "Dedicated Analytics (/crm/analytics)",
         "Verify Date Range Filter on Analytics page (Today, Last 7 Days, This Month, Custom Range)",
         "1. Select Date Filter -> 'This Month' on Analytics page\n2. Observe KPI cards and charts\n3. Switch to 'Custom Range'",
         "Date Range: Start Date to End Date", "Positive", "All analytics graphs, funnel data, and counselor stats should dynamically recalculate for the selected timeframe only."),
        
        ("TC_CRM_052", "CRM", "Analytics Page", "SC_CRM_10", "Dedicated Analytics (/crm/analytics)",
         "Verify exporting Analytics summary and Counsellor report to PDF/Excel",
         "1. On CRM Analytics page click 'Export Report' button\n2. Select format (PDF/Excel)\n3. Download and verify contents",
         "Report Export request", "Positive", "Downloaded report should format cleanly with charts/tables preserved and numbers matching on-screen analytics."),

        # 10. Mobile & Tablet Responsiveness
        ("TC_CRM_053", "CRM", "UI & Responsiveness", "SC_CRM_11", "Mobile & Tablet Responsiveness",
         "Verify Kanban Board touch drag-and-drop and responsiveness on mobile viewports (375px width)",
         "1. Open browser DevTools\n2. Switch to Mobile View (e.g. iPhone 12 / 375px width)\n3. Navigate to Kanban\n4. Attempt touch drag-and-drop",
         "Mobile viewport 375x812", "Positive", "Kanban columns should align horizontally with smooth touch swipe; dragging lead cards via touch should work without visual glitches."),
        
        ("TC_CRM_054", "CRM", "UI & Responsiveness", "SC_CRM_11", "Mobile & Tablet Responsiveness",
         "Verify Leads List table responsiveness on tablet and mobile screens",
         "1. Switch DevTools to Tablet (768px) and Mobile (375px) viewports\n2. Navigate to Leads List\n3. Check table alignment and scroll",
         "Tablet and Mobile viewports", "Positive", "Table should either adapt into responsive cards or provide a smooth horizontal scrollbar without breaking header alignment."),
        
        ("TC_CRM_055", "CRM", "UI & Responsiveness", "SC_CRM_11", "Mobile & Tablet Responsiveness",
         "Verify Quick Action buttons on mobile devices trigger native handlers (Dialer, WhatsApp app, Mail)",
         "1. On mobile device/viewport click Phone icon on lead card\n2. Click WhatsApp icon\n3. Click Email icon",
         "Lead with valid phone & email", "Positive", "Phone icon triggers native dialer; WhatsApp icon opens native WhatsApp app or wa.me link; Email opens default mail client."),

        # 11. Performance, Rate Limiting & Concurrency
        ("TC_CRM_056", "CRM", "Performance & Security", "SC_CRM_12", "Rate Limiting & High Load",
         "Verify API Rate Limiting protection against rapid double-clicking or spamming save/stage requests",
         "1. Open Add Lead modal or Kanban board\n2. Rapidly click 'Save' button 5 times within 1 second or send burst API requests",
         "Burst of 5 requests in 1 second", "Negative", "System should debounce button clicks or return HTTP 429 'Too Many Requests' to prevent duplicate lead creation or race conditions."),
        
        ("TC_CRM_057", "CRM", "Performance & Security", "SC_CRM_12", "Rate Limiting & High Load",
         "Verify Leads List and Kanban Board rendering performance with 1000+ active lead records",
         "1. Seed database with 1000+ test leads\n2. Open Leads List and Kanban Board\n3. Monitor page load time and UI responsiveness",
         "1000+ Lead Records in DB", "Positive", "Page should load within acceptable SLA (<2 seconds) using pagination or virtual scrolling without browser freezing."),

        # 12. Cross-Module Sync & Notifications
        ("TC_CRM_058", "CRM", "Cross-Module Sync", "SC_CRM_13", "ERP Module Integrations",
         "Verify course fee and batch data synchronization when converting Lead to Admission",
         "1. Convert a Lead interested in 'Data Science' course\n2. On Admission form verify if Course and Fee structure auto-populate from Course module",
         "Course: Data Science (Fee: Rs 50000)", "Positive", "Admission form should automatically fetch and select the exact Course, Batch options, and Fee structure linked to that course."),
        
        ("TC_CRM_059", "CRM", "Notifications", "SC_CRM_14", "Real-time Alerts & Reminders",
         "Verify in-app toast/alert notification when a scheduled follow-up time is reached",
         "1. Schedule a follow-up for 2 minutes from now\n2. Remain logged into ERP dashboard\n3. Wait for scheduled time",
         "Scheduled follow-up trigger time", "Positive", "An audio/visual notification toast or badge alert should pop up notifying the counselor that the follow-up is due right now."),
        
        ("TC_CRM_060", "CRM", "Performance & Security", "SC_CRM_12", "Rate Limiting & High Load",
         "Verify concurrent editing protection when two users modify the same lead simultaneously",
         "1. Login as Admin in Browser A and Counselor in Browser B\n2. Both open Lead ID #101\n3. Browser A updates Name and saves\n4. Browser B tries saving different Phone",
         "Concurrent session edit on same Lead ID", "Negative", "System should handle concurrency gracefully (e.g., last-write-wins with version check or notify User B that record was modified by another user)."),

        # 13. Sources & Stages Config
        ("TC_CRM_061", "CRM", "Sources & Stages Config", "SC_CRM_15", "Admin Configuration APIs (/lead-sources & stages)",
         "Verify adding a new custom Lead Source from Admin Settings",
         "1. Log in as Admin\n2. Navigate to CRM Settings / Lead Sources\n3. Click 'Add Source'\n4. Enter Name (e.g. 'LinkedIn Ads') and activate\n5. Save",
         "Source Name: LinkedIn Ads, Status: Active", "Positive", "New Lead Source should be saved and immediately appear in the 'Source' dropdown during new lead creation."),
        
        ("TC_CRM_062", "CRM", "Sources & Stages Config", "SC_CRM_15", "Admin Configuration APIs (/lead-sources & stages)",
         "Verify deactivating an obsolete Lead Source",
         "1. In Lead Sources list locate an old source (e.g. 'Old Newspaper')\n2. Toggle status to Inactive\n3. Open Add New Lead modal",
         "Source: Old Newspaper (Inactive)", "Positive", "Deactivated source should no longer appear in the new lead creation dropdown while preserving historical reports."),
        
        ("TC_CRM_063", "CRM", "Sources & Stages Config", "SC_CRM_15", "Admin Configuration APIs (/lead-sources & stages)",
         "Verify customizing Lead Stages (Adding a custom stage with hex color and sequence order)",
         "1. Navigate to Lead Stages config\n2. Click 'Add Stage'\n3. Enter Name ('Fee Negotiation'), Order Sequence (#4), Color Code ('#FF5733')\n4. Save",
         "Stage: Fee Negotiation, Order: 4, Color: #FF5733", "Positive", "New stage should appear on the Kanban Board at exact order #4 with header styled in custom hex color #FF5733."),

        # 14. Bulk Communication
        ("TC_CRM_064", "CRM", "Bulk Communication", "SC_CRM_16", "Bulk SMS & Email Broadcast (/communications)",
         "Verify sending a Bulk SMS campaign to selected leads via SMS Gateway (Twilio/MSG91)",
         "1. Navigate to CRM > Leads list\n2. Check 20 leads\n3. Click 'Bulk Action -> Send SMS'\n4. Select SMS template or type custom text\n5. Execute Send",
         "20 Selected Leads; SMS Text: 'Admission open for weekend batch!'", "Positive", "System should queue and dispatch SMS via Twilio/MSG91 gateway and show success summary 'SMS queued for 20 recipients'."),
        
        ("TC_CRM_065", "CRM", "Bulk Communication", "SC_CRM_16", "Bulk SMS & Email Broadcast (/communications)",
         "Verify sending an HTML formatted Bulk Email broadcast to filtered leads",
         "1. Filter leads by Stage = 'Contacted'\n2. Click 'Bulk Action -> Send Email'\n3. Enter Subject and HTML Body\n4. Execute Send",
         "Subject: 'Exclusive Diwali Discount on Courses'; HTML Body", "Positive", "Email broadcast should be sent via SendGrid/SMTP without formatting loss or SMTP timeout errors."),
        
        ("TC_CRM_066", "CRM", "Bulk Communication", "SC_CRM_16", "Bulk SMS & Email Broadcast (/communications)",
         "Verify checking SMS and Email Delivery Logs (/sms-logs & /email-logs)",
         "1. Navigate to CRM > Communications > SMS / Email Logs\n2. Review recent broadcast entries\n3. Check status badges and gateway responses",
         "Historical broadcast records", "Positive", "Logs should display recipient details, timestamp, message content, and exact gateway status ('Sent', 'Delivered', 'Failed')."),
        
        ("TC_CRM_067", "CRM", "Bulk Communication", "SC_CRM_16", "Bulk SMS & Email Broadcast (/communications)",
         "Verify validation when initiating Bulk Communication without selecting recipients or with blank body",
         "1. Click 'Send Bulk SMS/Email' with 0 leads selected OR leave message body blank\n2. Attempt sending",
         "Recipients: 0 OR Message: (blank)", "Negative", "System should block submission and display error 'Please select at least 1 recipient' or 'Message body cannot be empty'."),

        # 15. External Integrations (Webhooks)
        ("TC_CRM_068", "CRM", "External Integrations", "SC_CRM_17", "Webhook Lead Intake (Google/FB/JustDial)",
         "Verify automatic lead creation via Google Ads Lead Form Extension Webhook (/leads/google)",
         "1. Use Postman/curl to simulate POST webhook payload to /api/v1/leads/google with sample lead data\n2. Check CRM Leads list",
         "Google Ads JSON Payload (Name, Phone, Email, Campaign)", "Positive", "System should intercept webhook, validate payload, auto-create a new lead with source tagged as 'Google Ads', and notify counselor."),
        
        ("TC_CRM_069", "CRM", "External Integrations", "SC_CRM_17", "Webhook Lead Intake (Google/FB/JustDial)",
         "Verify automatic lead capture via Facebook Lead Ads API Webhook (/leads/facebook)",
         "1. Simulate Facebook Lead Ads POST webhook to /api/v1/leads/facebook\n2. Verify lead creation and phone formatting",
         "Facebook Lead JSON Payload", "Positive", "Lead should be created instantaneously with accurate field mapping and source set to 'Facebook Lead Ads'."),
        
        ("TC_CRM_070", "CRM", "External Integrations", "SC_CRM_17", "Webhook Lead Intake (Google/FB/JustDial)",
         "Verify Just Dial API webhook lead intake and deduplication (/leads/justdial)",
         "1. Send POST request to /api/v1/leads/justdial with phone number already existing in DB\n2. Observe response and DB state",
         "Just Dial Payload with duplicate phone", "Negative", "System should identify duplicate phone number, log the Just Dial enquiry under existing lead's Activities Timeline instead of creating duplicate."),
        
        ("TC_CRM_071", "CRM", "External Integrations", "SC_CRM_17", "Webhook Lead Intake (Google/FB/JustDial)",
         "Verify webhook security against unauthorized POST requests (Invalid API Secret/Signature)",
         "1. Send POST request to /api/v1/leads/google without valid X-Hub-Signature or API Secret token\n2. Check response",
         "Unauthorized JSON Payload without header token", "Negative", "API Gateway should reject request with HTTP 401 Unauthorized or 403 Forbidden and prevent lead creation."),

        # 16. Auto Assignment Engine
        ("TC_CRM_072", "CRM", "Auto Assignment Engine", "SC_CRM_18", "Lead Routing & Assignment Rules Engine",
         "Verify Round-Robin Auto Assignment rule distributes incoming leads equally among active counselors",
         "1. Configure Auto Assignment Rule = 'Round Robin' for Counselors A, B, and C\n2. Create 3 new leads sequentially via API/form\n3. Check assigned_to field",
         "3 New Incoming Leads", "Positive", "Lead 1 assigns to Counselor A; Lead 2 assigns to Counselor B; Lead 3 assigns to Counselor C automatically."),
        
        ("TC_CRM_073", "CRM", "Auto Assignment Engine", "SC_CRM_18", "Lead Routing & Assignment Rules Engine",
         "Verify Course-Based Auto Assignment routing rule",
         "1. Configure rule: If Course Interest = 'Data Science', assign to Senior Counselor X\n2. Create lead with course 'Data Science'\n3. Check assignment",
         "Course: Data Science", "Positive", "Lead should bypass round-robin and assign directly to Senior Counselor X with automatic notification sent to Counselor X."),
        
        ("TC_CRM_074", "CRM", "Auto Assignment Engine", "SC_CRM_18", "Lead Routing & Assignment Rules Engine",
         "Verify Manual Override reassignment by Branch Admin / Team Leader",
         "1. Log in as Branch Admin\n2. Open a lead assigned to Counselor A\n3. Click 'Reassign' and select Counselor B\n4. Save",
         "Reassign from Counselor A to Counselor B", "Positive", "Lead ownership should update immediately to Counselor B; audit log should record reassignment by Admin; Counselor B receives alert."),
        
        ("TC_CRM_075", "CRM", "Analytics Page", "SC_CRM_10", "Dedicated Analytics (/crm/analytics)",
         "Verify Course-Wise Enquiry & Conversion Report (/analytics/course-wise)",
         "1. Navigate to CRM > Analytics > Course-Wise Report\n2. Inspect table/chart showing enquiries per course\n3. Check conversion % column",
         "Multi-course enquiry data", "Positive", "Report should accurately display total enquiries, converted admissions, and conversion rate percentage categorized by each course.")
    ]

    # Map raw tuples into strict 17-column dictionaries
    dict_cases = []
    for c in raw_cases:
        item = {
            "Test Case ID": c[0],
            "Module": c[1],
            "Sub Module": c[2],
            "Scenario ID": c[3],
            "Scenario description": c[4],
            "Tester Name": "",
            "Date": "",
            "Test case description": c[5],
            "Test Steps": c[6],
            "Test Data": c[7],
            "Type of Test cases": c[8],
            "Expected results": c[9],
            "Date of execution": "",
            "Executed By": "",
            "Test Status": "Not Executed",
            "Actual Result": "",
            "Defect ID": ""
        }
        dict_cases.append(item)

    # 1. Write clean CSV
    csv_file = r"d:\webapp\CRM_Test_Cases_Detailed.csv"
    try:
        with open(csv_file, mode="w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            for d in dict_cases:
                writer.writerow(d)
        print(f"Successfully generated clean CSV without column shifts: {csv_file}")
    except PermissionError:
        alt_csv = r"d:\webapp\CRM_Test_Cases_Detailed_v2.csv"
        with open(alt_csv, mode="w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            for d in dict_cases:
                writer.writerow(d)
        print(f"Successfully generated clean CSV: {alt_csv} (Original open in Excel)")

    # 2. Write styled Excel XLSX
    xlsx_file = r"d:\webapp\CRM_Test_Cases_Detailed.xlsx"
    alt_xlsx = r"d:\webapp\CRM_Test_Cases_Detailed_v2.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRM Test Cases"

    # Write header
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    for row_idx, d in enumerate(dict_cases, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=d[col_name])

    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    header_border = Border(
        left=Side(style='thin', color='1F4E78'), right=Side(style='thin', color='1F4E78'),
        top=Side(style='thin', color='1F4E78'), bottom=Side(style='medium', color='000000')
    )

    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    left_align = Alignment(horizontal="left", vertical="top")

    col_settings = {
        1:  {"width": 15, "align": center_align},  # Test Case ID
        2:  {"width": 14, "align": center_align},  # Module
        3:  {"width": 24, "align": left_align},    # Sub Module
        4:  {"width": 14, "align": center_align},  # Scenario ID
        5:  {"width": 26, "align": wrap_align},    # Scenario description
        6:  {"width": 15, "align": left_align},    # Tester Name
        7:  {"width": 12, "align": center_align},  # Date
        8:  {"width": 32, "align": wrap_align},    # Test case description
        9:  {"width": 45, "align": wrap_align},    # Test Steps
        10: {"width": 28, "align": wrap_align},    # Test Data
        11: {"width": 18, "align": center_align},  # Type of Test cases
        12: {"width": 45, "align": wrap_align},    # Expected results
        13: {"width": 16, "align": center_align},  # Date of execution
        14: {"width": 15, "align": left_align},    # Executed By
        15: {"width": 15, "align": center_align},  # Test Status
        16: {"width": 25, "align": wrap_align},    # Actual Result
        17: {"width": 12, "align": center_align}   # Defect ID
    }

    max_col = len(columns)
    max_row = len(dict_cases) + 1

    # Header styling
    for col_idx in range(1, max_col + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = header_border
    ws.row_dimensions[1].height = 28

    # Data styling
    for row_idx in range(2, max_row + 1):
        zebra_fill = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid") if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx in range(1, max_col + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = data_font
            c.border = thin_border
            if zebra_fill.fill_type:
                c.fill = zebra_fill
            if col_idx in col_settings:
                c.alignment = col_settings[col_idx]["align"]
            else:
                c.alignment = left_align

    # Column widths
    for col_idx, settings in col_settings.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = settings["width"]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    try:
        wb.save(xlsx_file)
        print(f"Successfully generated perfect Excel workbook: {xlsx_file}")
    except PermissionError:
        wb.save(alt_xlsx)
        print(f"Successfully generated perfect Excel workbook: {alt_xlsx} (Original open in Excel)")

if __name__ == "__main__":
    build_test_cases()
