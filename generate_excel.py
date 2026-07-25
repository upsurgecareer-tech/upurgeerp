import csv
import sys
import subprocess
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def create_excel():
    csv_file = r"d:\webapp\CRM_Test_Cases_Detailed.csv"
    xlsx_file = r"d:\webapp\CRM_Test_Cases_Detailed.xlsx"

    if not os.path.exists(csv_file):
        print(f"Error: {csv_file} not found!")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRM Test Cases"

    # Read CSV and write to Worksheet
    with open(csv_file, mode='r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Professional Dark Slate Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    header_border = Border(
        left=Side(style='thin', color='1F4E78'),
        right=Side(style='thin', color='1F4E78'),
        top=Side(style='thin', color='1F4E78'),
        bottom=Side(style='medium', color='000000')
    )

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(horizontal="center", vertical="top")
    left_alignment = Alignment(horizontal="left", vertical="top")

    # Column widths and alignment settings
    col_settings = {
        1:  {"width": 15, "align": center_alignment},  # Test Case ID
        2:  {"width": 16, "align": center_alignment},  # Module
        3:  {"width": 24, "align": left_alignment},    # Sub Module
        4:  {"width": 14, "align": center_alignment},  # Scenario ID
        5:  {"width": 28, "align": wrap_alignment},    # Scenario description
        6:  {"width": 15, "align": left_alignment},    # Tester Name
        7:  {"width": 12, "align": center_alignment},  # Date
        8:  {"width": 28, "align": wrap_alignment},    # Test case description
        9:  {"width": 45, "align": wrap_alignment},    # Test Steps
        10: {"width": 30, "align": wrap_alignment},    # Test Data
        11: {"width": 18, "align": center_alignment},  # Type of Test cases
        12: {"width": 45, "align": wrap_alignment},    # Expected results
        13: {"width": 15, "align": center_alignment},  # Date of execution
        14: {"width": 15, "align": left_alignment},    # Executed By
        15: {"width": 15, "align": center_alignment},  # Test Status
        16: {"width": 25, "align": wrap_alignment},    # Actual Result
        17: {"width": 12, "align": center_alignment}   # Defect ID
    }

    max_row = ws.max_row
    max_col = ws.max_column

    # Apply Header Styling
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    ws.row_dimensions[1].height = 28  # Taller header row

    # Apply Data Styling
    for row_idx in range(2, max_row + 1):
        # Optional: Alternating row color (zebra striping) for readability
        zebra_fill = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid") if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if zebra_fill.fill_type:
                cell.fill = zebra_fill
            
            # Apply alignment based on column
            if col_idx in col_settings:
                cell.alignment = col_settings[col_idx]["align"]
            else:
                cell.alignment = left_alignment

    # Set exact column widths
    for col_idx, settings in col_settings.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = settings["width"]

    # Freeze Top Header Row
    ws.freeze_panes = "A2"

    # Enable Auto-Filter on headers
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # Save Workbook
    wb.save(xlsx_file)
    print(f"Successfully generated professional Excel workbook: {xlsx_file}")

if __name__ == "__main__":
    create_excel()
